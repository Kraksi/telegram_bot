import openpyxl
import os
import pathlib
from data_base import sqlite_db
import pyexcel
import io


def openxlsx():
    wb = openpyxl.load_workbook(filename='file.xlsx', read_only=True)
    ws = wb['Лист1']
    for row in ws.rows:
        buf_list = []
        for cell in row:
            buf_list.append(cell.value)
        print(buf_list)
    wb.close()
    os.remove('C:/Users/Krasti/Desktop/telegram_bot/file.xlsx')
    #   xlsx_filename = "file.xlsx"
    #    with open(xlsx_filename, "rb") as f: in_mem_file = io.BytesIO(f.read())
    #    wb = openpyxl.load_workbook(in_mem_file, read_only=True)
    #    wb = openpyxl.load_workbook(in_mem_file, read_only= False)
    #    worksheet = wb.active
    #    for i in range(0, worksheet.max_row):
    #        buf_list = []
    #        for col in worksheet.iter_cols(1, worksheet.max_column):
    #            buf_list.append(col[i].value)
    #        sqlite_db.add_many_q(buf_list)
    print("success in function")
